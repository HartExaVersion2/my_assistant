from openpyxl import load_workbook
from common.errors import InitError, UnknownError

class FillTable():

    def __init__(self, filename, logger):
        try:
            self.filename = filename
            self.wb = load_workbook(self.filename)
            self.sheet = self.wb.active
            self.logger = logger
        except Exception as error:
            message = 'class FillTable error {}'.format(error)
            raise InitError(message)

    def __searching_empty_column(self) -> int:
        try:
            for i in range(1, 40):
                for row in self.sheet.iter_rows(min_row=2, max_row=13, min_col=i, max_col=i, values_only=False):
                    column_value = row[0].value
                    if column_value == '-':
                        return i
            return -1
        except Exception as error:
            message = 'function {} error {}'.format(self.__searching_empty_column.__name__, error)
            self.logger.error(message)
            raise UnknownError

    def fill_table(self, dict_count_stars: dict):
        try:
            empty_column = self.__searching_empty_column()
            for row in self.sheet.iter_rows(min_row=5, max_row=12, min_col=2, max_col=2, values_only=False):
                surname = row[0].value  # Значение фамилии в текущей строке
                if surname in dict_count_stars:
                    value = dict_count_stars[surname]  # Получаем значение из словаря по фамилии
                    self.sheet.cell(row=row[0].row, column=empty_column).value = value
                else:
                    self.sheet.cell(row=row[0].row, column=empty_column).value = ''
            self.wb.save(self.filename)
        except Exception as error:
            message = 'function {} error {}'.format(self.fill_table.__name__, error)
            self.logger.error(message)
            raise UnknownError